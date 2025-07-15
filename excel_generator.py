"""
Excel report generation functions
"""
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import io
from datetime import datetime


def create_excel_report(data, today=None):
    """Create formatted Excel report with all sheets"""
    if today is None:
        today = datetime.now()
    
    # Create Excel file in memory
    output = io.BytesIO()
    
    # Write all sheets first
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        write_all_sheets(writer, data)

    # Format all sheets
    output.seek(0)
    wb = openpyxl.load_workbook(output)
    format_all_sheets(wb, today)

    # Save to bytes
    output_formatted = io.BytesIO()
    wb.save(output_formatted)
    output_formatted.seek(0)
    
    return output_formatted.getvalue()


def write_all_sheets(writer, data):
    """Write data to all Excel sheets"""
    data['processed_df'].to_excel(writer, sheet_name="Referral Overview", index=False)
    data['summary'].to_excel(writer, sheet_name="Pending Tasks Summary", index=False, startrow=1)
    data['cchp_nutrition'].to_excel(writer, sheet_name="Pending CCHP Nutrition", index=False)
    data['initial_mtg'].to_excel(writer, sheet_name="Pending Initial MTG Box", index=False)
    data['ongoing_mtg'].to_excel(writer, sheet_name="Pending Ongoing MTG Box", index=False)
    data['nutritional_assessment'].to_excel(writer, sheet_name="Pending Nutrition Assess", index=False)
    data['speak_to_member'].to_excel(writer, sheet_name="Pending Speak to Member", index=False)
    data['tar_approval'].to_excel(writer, sheet_name="Pending TAR Approval", index=False)
    data['reauth_pending'].to_excel(writer, sheet_name="Pending Reauth NotSubm", index=False)


def format_all_sheets(wb, today):
    """Apply formatting to all sheets in the workbook"""
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    info_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    bold_font = Font(bold=True)

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        format_sheet(ws, sheetname, today, header_fill, info_fill, bold_font)


def format_sheet(ws, sheetname, today, header_fill, info_fill, bold_font):
    """Format a single worksheet"""
    ws.freeze_panes = "A2"
    
    # Only enable autofilter for non-summary sheets
    if sheetname != "Pending Tasks Summary":
        ws.auto_filter.ref = ws.dimensions
        header_row = 1
    else:
        header_row = format_summary_sheet(ws, today, info_fill, bold_font, header_fill)

    # Format headers and auto-adjust column widths
    format_headers_and_columns(ws, header_row, header_fill, bold_font)


def format_summary_sheet(ws, today, info_fill, bold_font, header_fill):
    """Special formatting for the summary sheet"""
    ws.insert_rows(1)
    ws["A1"] = f"Data is based on: {today.strftime('%Y-%m-%d %I:%M %p')}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    
    # Format timestamp row
    for cell in ws[1]:
        cell.fill = info_fill
        cell.font = bold_font
    
    # Format header row
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = bold_font
    
    # Clear formatting for empty row
    for cell in ws[2]:
        cell.fill = PatternFill(fill_type=None)
    
    return 3  # Header row number


def format_headers_and_columns(ws, header_row, header_fill, bold_font):
    """Format headers and adjust column widths"""
    for col_idx, cell in enumerate(ws[header_row], start=1):
        cell.fill = header_fill
        cell.font = bold_font
        
        # Auto-adjust column width
        max_length = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in ws[get_column_letter(col_idx)]
        )
        adjusted_width = max_length + 2
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = adjusted_width
