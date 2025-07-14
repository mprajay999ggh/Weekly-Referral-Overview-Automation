import pandas as pd
from datetime import timedelta
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl import load_workbook

# === Load Data ===
file_path = "Umoja Referral Overview 0617 BPH.xlsx"  
df = pd.read_excel(file_path, engine='openpyxl', keep_default_na=False)

print(df.head())
today = pd.to_datetime("today").normalize()
#today = pd.to_datetime("2025-06-18").normalize()
#print(today - timedelta(days=49))
# === Clean & Convert Columns ===
df['Referral Start Date'] = pd.to_datetime(df['Referral Start Date'], errors='coerce')
df['Day(s) in Current Activity'] = (today - df['Last Activity Date']).dt.days
#df['Day(s) in Current Activity'] = pd.to_numeric(df['Day(s) in Current Activity'], errors='coerce')
df['Number of Grocery Boxes Successfully Sent'] = pd.to_numeric(df['Number of Grocery Boxes Successfully Sent'], errors='coerce')
df['Number of Nutrition Counseling Sessions Completed'] = pd.to_numeric(df['Number of Nutrition Counseling Sessions Completed'], errors='coerce')

print(today - timedelta(days=49))

# === Task Logic ===
initial_mtg = df[
    (df['Pending Task/ Next Task'] == "MTG Box Delivery") &
    (df['Day(s) in Current Activity'] >= 4) &
    (df['Number of Grocery Boxes Successfully Sent'] == 0)
]

ongoing_mtg = df[
    (df['Pending Task/ Next Task'] == "MTG Box Delivery") &
    (df['Day(s) in Current Activity'] >= 8) &
    (df['Number of Grocery Boxes Successfully Sent'] != 0)
]

nutritional_assessment = df[
    (df['Pending Task/ Next Task'] == "Nutritional assessment") &
    (df['Day(s) in Current Activity'] >= 14)
]

speak_to_member = df[
    (df['Pending Task/ Next Task'] == "Speak to Member") &
    (df['Day(s) in Current Activity'] >= 14)
]

tar_approval = df[
    (df['Pending Task/ Next Task'] == "TAR Approval") &
    (df['Day(s) in Current Activity'] >= 8)
]

cchp_nutrition = df[
    (df['Payer Organization'].str.upper() == "CCHP") &
    (df['Referral Created Date'] <= today - timedelta(days=49)) &
    (df['Number of Nutrition Counseling Sessions Completed'].isin([0,1])) &
    (~df['Pending Task/ Next Task'].astype(str).str.lower().str.contains("discontinued"))
]


def is_reauth_due(row):
    try:
        if str(row['Re-authorization Status']).strip().upper() != "NA":
            return False
        if str(row['Pending Task/ Next Task']).lower() in ["services discontinued", "service discontinued"]:
            return False
        if str(row.get('Last Activity Completed')).strip().lower() == "reauthorization approved":
            return False

        start_date = row['Referral Start Date']
        if pd.isnull(start_date):
            return False

        org = str(row['Payer Organization']).strip().upper()
        if org == "CCHP":
            return today >= start_date + timedelta(weeks=11)
        elif org == "CCAH":
            return today >= start_date + timedelta(weeks=15)
        elif org == "PHP":
            return today >= pd.to_datetime(start_date) + pd.DateOffset(months=5)
    except:
        return False
    return False


reauth_pending = df[df.apply(is_reauth_due, axis=1)]

# === Summary Table ===
summary = pd.DataFrame({
    "Category": [
        "INITIAL MTG box delivery",
        "ONGOING MTG box delivery",
        "Nutritional assessment",
        "Speak to member",
        "TAR approval",
        "Nutritional counseling",
        "Reauth not submitted"
    ],
    "Number of Referrals": [
        len(initial_mtg),
        len(ongoing_mtg),
        len(nutritional_assessment),
        len(speak_to_member),
        len(tar_approval),
        len(cchp_nutrition),
        len(reauth_pending)
    ],
    "Definition": [
        "4 or more days pending delivery of initial box",
        "8 or more days pending delivery of follow-up boxes",
        "14 or more days pending nutritional assessment",
        "14 or more days pending speak to member status",
        "8 or more days pending TAR approval",
        "9 weeks from referral start date for CCHP",
        "CCHP - 11 weeks (out of 12)\nCCAH - 15 weeks (out of 17)\nPHP - 5 months (out of 6)"
    ]
})



# === Save to Excel with separate sheets for each metric ===
output_path = "referral_dashboard.xlsx"

# Write all sheets first, then format after closing ExcelWriter
import openpyxl
from datetime import datetime
now = datetime.now()
with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Referral Overview", index=False)
    summary.to_excel(writer, sheet_name="Pending Tasks Summary", index=False, startrow=1)
    cchp_nutrition.to_excel(writer, sheet_name="Pending CCHP Nutrition", index=False)
    initial_mtg.to_excel(writer, sheet_name="Pending Initial MTG Box", index=False)
    ongoing_mtg.to_excel(writer, sheet_name="Pending Ongoing MTG Box", index=False)
    nutritional_assessment.to_excel(writer, sheet_name="Pending Nutrition Assess", index=False)
    speak_to_member.to_excel(writer, sheet_name="Pending Speak to Member", index=False)
    tar_approval.to_excel(writer, sheet_name="Pending TAR Approval", index=False)
    reauth_pending.to_excel(writer, sheet_name="Pending Reauth NotSubm", index=False)



# Now format all sheets, handling Pending Tasks Summary specially
wb = openpyxl.load_workbook(output_path)
header_fill = openpyxl.styles.PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
info_fill = openpyxl.styles.PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
bold_font = openpyxl.styles.Font(bold=True)

for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    ws.freeze_panes = "A2"
    # Only enable autofilter for non-summary sheets
    if sheetname != "Pending Tasks Summary":
        ws.auto_filter.ref = ws.dimensions
    if sheetname == "Pending Tasks Summary":
        ws.insert_rows(1)
        ws["A1"] = f"Data is based on: {now.strftime('%Y-%m-%d %I:%M %p')}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
        for cell in ws[1]:
            cell.fill = info_fill
            cell.font = bold_font
        for cell in ws[3]:
            cell.fill = header_fill
            cell.font = bold_font
        for cell in ws[2]:
            cell.fill = openpyxl.styles.PatternFill(fill_type=None)
        header_row = 3
    else:
        header_row = 1
    for col_idx, cell in enumerate(ws[header_row], start=1):
        cell.fill = header_fill
        cell.font = bold_font
        max_length = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in ws[get_column_letter(col_idx)]
        )
        adjusted_width = max_length + 2
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = adjusted_width

wb.save(output_path)
print(f"✅ Beautified dashboard saved to: {output_path} (with separate sheets for each metric)")
